import random
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font
import msoffcrypto
from file_generators.base.base_file import BaseFile
from file_generators.utils.helpers import generate_random_id


class ExcelFile(BaseFile):
    """Excel file generator with malware signature preservation support."""

    def get_file_extension(self) -> str:
        return ".xlsx"

    def generate(self, data: str, location: str = "A1",
                 encrypt: bool = False, preserve_content: bool = False, **kwargs) -> bytes:
        """
        Generate Excel file with data.

        Args:
            data: Text content for cell
            location: Cell location (e.g., "A1", "B2")
            encrypt: Whether to encrypt the file
            preserve_content: If True, preserves content exactly (critical for EICAR)
            **kwargs: Additional parameters (ignored)

        Returns:
            Excel file bytes
        """
        try:
            if preserve_content:
                print("ExcelFile: Preserving content exactly (EICAR mode)")
                # CRITICAL: For EICAR, use exact content - NO random ID
                full_data = data
                use_random_color = False
            else:
                print(f"ExcelFile: Standard mode, location={location}")
                # Standard mode: add random ID
                full_data = data + "\n" + generate_random_id()
                use_random_color = True

            # Create workbook
            wb = Workbook()
            ws = wb.active

            # Insert data
            cell = ws[location] if location else ws['A1']
            cell.value = full_data

            # Set font color based on mode
            if use_random_color:
                # Standard mode: random color
                random_color = "{:02X}{:02X}{:02X}".format(
                    random.randint(0, 255),
                    random.randint(0, 255),
                    random.randint(0, 255)
                )
                cell.font = Font(color=random_color)
            else:
                # Preserved mode: use default/black color
                cell.font = Font(color="000000")  # Black

            # Save to bytes
            byte_io = BytesIO()
            wb.save(byte_io)
            byte_io.seek(0)

            if encrypt:
                print(f"ExcelFile: Encrypting Excel file with password")
                encrypted_bytes = self._encrypt_workbook(byte_io)
                if preserve_content:
                    print(f"ExcelFile: Generated encrypted Excel with preserved content ({len(encrypted_bytes)} bytes)")
                return encrypted_bytes
            else:
                byte_data = byte_io.getvalue()
                if preserve_content:
                    print(f"ExcelFile: Generated Excel with preserved content ({len(byte_data)} bytes)")
                else:
                    print(f"ExcelFile: Generated standard Excel at {location} ({len(byte_data)} bytes)")
                return byte_data

        except Exception as e:
            print(f"Error generating Excel file: {e}")
            raise

    def _encrypt_workbook(self, byte_io: BytesIO) -> bytes:
        """Encrypt Excel workbook with password."""
        try:
            office_file = msoffcrypto.OfficeFile(byte_io)
            encrypted_buffer = BytesIO()
            office_file.encrypt(self.password, encrypted_buffer)
            encrypted_buffer.seek(0)
            return encrypted_buffer.getvalue()
        except Exception as e:
            print(f"Error encrypting Excel workbook: {e}")
            raise
